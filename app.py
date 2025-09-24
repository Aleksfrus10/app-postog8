import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Lista de campos
campos = [
    "VIGILANTE LÍDER (UN-30)",
    "RP-23",
    "RP-22",
    "RP-21",
    "PORTARIA 01 A",
    "PORTARIA 01 B",
    "PORTARIA 01 C",
    "PORTARIA 02",
    "PORTARIA 04",
    "PORTARIA 06",
    "PORTARIA 07",
    "PORTARIA 09",
    "PORTARIA 10 A",
    "PORTARIA 10 B",
    "PORTARIA 10 C",
    "POSTO 11",
    "PORTARIA 15 A",
    "PORTARIA 15 B",
    "PORTARIA 16 A",
    "PORTARIA 16 B",
    "TUBOVIA",
    "UTE/S-10",
    "CRUZEIRO",
    "CFTV-UTE",
    "CISP A",
    "CISP B",
    "RECEPÇÃO PV-10",
    "CREDENCIAMENTO",
    "RECEPÇÃO CEAD A",
    "RECEPÇÃO CEAD B",
    "RECEPÇÃO CEAD UTE",
    "RECEPÇÃO BALANÇA"
]

st.set_page_config(page_title="Formulário de Postos", layout="wide")
st.title("Formulário de Preenchimento de Postos")
st.write("Preencha os campos abaixo e clique em 'Enviar'. O arquivo Excel será gerado para download.")

# Formulário em Streamlit
with st.form("form_postos"):
    dados = {}
    for campo in campos:
        dados[campo] = st.text_input(campo)
    
    enviar = st.form_submit_button("Enviar")

    if enviar:
        # Validação: campos vazios
        campos_vazios = [c for c, v in dados.items() if not v.strip()]
        if campos_vazios:
            st.error(f"Os seguintes campos não podem ficar vazios: {', '.join(campos_vazios)}")
        else:
            # Criar DataFrame transposto
            df = pd.DataFrame([dados]).T
            df.columns = ['Preenchimento']

            # Exibir DataFrame
            st.success("Formulário enviado com sucesso!")
            st.dataframe(df)

            # Criar arquivo Excel em memória
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=True, sheet_name='Postos')
                ws = writer.sheets['Postos']

                # Estilo básico
                thin_border = Border(left=Side(style='thin'),
                                     right=Side(style='thin'),
                                     top=Side(style='thin'),
                                     bottom=Side(style='thin'))

                # Formatar cabeçalho
                for cell in ws[1]:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill("solid", fgColor="4F81BD")  # azul
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = thin_border

                # Ajustar largura das colunas
                for col in ws.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    ws.column_dimensions[column].width = max_length + 5

                # Centralizar valores e aplicar cor alternada nas linhas
                for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=2), start=2):
                    fill_color = "DCE6F1" if i % 2 == 0 else "FFFFFF"
                    for cell in row:
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        cell.fill = PatternFill("solid", fgColor=fill_color)
                        cell.border = thin_border

            output.seek(0)

            # Botão de download
            st.download_button(
                label="Baixar Excel",
                data=output,
                file_name="preenchimento_postos.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
